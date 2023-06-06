from django.core.management import BaseCommand
from openpyxl import load_workbook

from mysite.models import MySite


class Command(BaseCommand):
    def handle(self, *args, **options):
        filename = 'mysite.xlsx'
        wb = load_workbook(filename, read_only=True)
        ws = wb.active
        name, phone_number = list(ws.iter_rows(min_row=2, values_only=True))[0]
        MySite.objects.update_or_create(
            name=name,
            defaults={
                'phone_number': phone_number,
            },
        )


